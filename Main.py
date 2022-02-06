import json
import urllib
from urllib.parse import urlparse
import httplib2 as http #External library
import time 
from openpyxl import load_workbook

headers = { 'AccountKey' : 'Ynkrfx+oTXOQkNEvXcLfZg==','accept' : 'application/json'} #this is by default
#API parameters
uri = 'http://datamall2.mytransport.sg/' #Resource URL
path = 'ltaodataservice/BusArrivalv2?BusStopCode=21581&ServiceNo=240'
#Build query string & specify type of API call
target = urlparse(uri + path)
# print(target.geturl())
# ^ will print API URL 
method = 'GET'
body = ''
 #Get handle to http
h = http.Http()
#Obtain results
response, content = h.request(target.geturl(),method,body,headers)
 #Parse JSON to print
jsonObj = json.loads(content)
# print(json.dumps(jsonObj, sort_keys=True, indent=4))
# ^ will print readable schedule  

def get_current_time():
    t = time.localtime()
    current_time = time.strftime("%H:%M:%S",t)
    return current_time
    # current_time is current time

def next_bus_time_tuple():
    h = http.Http()
    response, content = h.request(target.geturl(),method,body,headers)
    jsonObj = json.loads(content)
    bus_service = jsonObj["Services"][0]
    next_bus_estimated_arrival = bus_service["NextBus"]["EstimatedArrival"][11:19]
    next_bus_estimated_arrival2 = bus_service["NextBus2"]["EstimatedArrival"][11:19]
    next_bus_estimated_arrival3 = bus_service["NextBus3"]["EstimatedArrival"][11:19]
    return (next_bus_estimated_arrival , next_bus_estimated_arrival2, next_bus_estimated_arrival3)

def time_difference_calculator(previous_time , new_time):
    previous_time = int(previous_time.split(":")[1]) * 60 + int(previous_time.split(":")[2])
    new_time = int(new_time.split(":")[1]) * 60 + int(new_time.split(":")[2])
    return new_time - previous_time
    #should always return positive 


def main():
    wb = load_workbook("Bus time.xlsx")
    ws = wb["240"] 
    starting_letter_chr = 66
    starting_letter = chr(starting_letter_chr) #corresponds to B 
    starting_number = 5
    print(f"Operation start, time is {get_current_time()}")
    print()

    #starting cell = B5
    cell = ws["B5"]
    cell.value = next_bus_time_tuple()[0]
    pbt = ws["B5"].value

    #assign expected time diff between busses
    expected_time_diff = time_difference_calculator(next_bus_time_tuple()[0],next_bus_time_tuple()[1])

    for i in range(60 * 9):
        nbt = next_bus_time_tuple()[0]
        nbt2 = next_bus_time_tuple()[1]
        actual_time_difference = time_difference_calculator(pbt,nbt)
        expected_minus_actual = expected_time_diff - actual_time_difference
        ct = get_current_time()
        cellname = starting_letter + str(starting_number)
        print(f"{i+1}.current time is {ct},next bus time is {nbt}")        
        
        if -60 <= expected_minus_actual <= 60:
        # bus has passed stop, start recording for next bus
            starting_number += 1 
            cell = ws[cellname]
            cell.value = nbt
            print("start new line")

        elif  nbt != pbt :
        # next bus havnt reach stop, timing different from previously updated
        # previous_bus_timing updated to match newest timing
            cell = ws[cellname]
            cell.value = nbt 
            print("update same cell")
    
        pbt = nbt #update pbt to newest time
        expected_time_diff = time_difference_calculator(next_bus_time_tuple()[0],next_bus_time_tuple()[1])
        print()
        time.sleep(60)

    wb.save("Bus time.xlsx")
    print("operation complete")

main()

# every min get nbt
# if != previous time, update cell 
# if nbt > current_time, means bus has passed stop, next bus cell shift down
